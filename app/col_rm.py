from flask import Blueprint, request, flash, redirect, render_template, send_from_directory, abort, send_file
from werkzeug.utils import secure_filename
from pathlib import Path
import openpyxl as oxl
import os

bp = Blueprint('colclear', __name__)

allowed_extensions = set(['txt', 'xlsx'])

path = "app/filesuploaded"
path1 = "app/col_rm_final"
# for pythonanywhere paths
# path = "/home/Flippy9004/excel_app/app/static/filesuploaded"
# path1 = "/home/Flippy9004/excel_app/app/static/col_rm_final"


def allowed_files(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


@bp.route('/clupload', methods=["GET", "POST"])
def clupload():
    # Check if there is any "final" file in directory for final files
    # If there is, remove it
    dirs = os.listdir(path1)
    for file in dirs:
        os.remove(path1 + '/' + file)

    if request.method == "POST":

        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)

        if 'file1' not in request.files:
            flash('No file part')
            return redirect(request.url)

        filse = request.files['file']
        filse1 = request.files['file1']

        files = [filse, filse1]
        for file in files:
            if file.filename == '':
                flash('No file selected')
                return redirect(request.url)

        for file in files:
            if file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
                files.remove(file)

        if len(files) == 1:
            fext = files[0].filename.split('.', 1)[1]
            if fext == 'txt':
                return render_template("col_rm/upload.html", numfiles=len(files), fext=True, wrong=True)
            else:
                return render_template("col_rm/upload.html", numfiles=len(files), fext=False, wrong=True)
        elif len(files) == 2:
            fext = files[0].filename.split('.', 1)[1]
            fext1 = files[1].filename.split('.', 1)[1]
            if fext == fext1 and fext == 'txt':
                return render_template("col_rm/upload.html", fext=True, wrong=True)
            elif fext == fext1 and fext == 'xlsx':
                return render_template("col_rm/upload.html", fext=False, wrong=True)

        filenames = []
        for file in files:
            if file and allowed_files(file.filename):
                filename = secure_filename(file.filename)
                filenames.append(filename)
                file.save(os.path.join(Path(path), filename))

        return render_template("col_rm/upload.html", filse=filenames, lfse=len(filenames), wrong=False)
    return render_template("col_rm/upload.html")


@bp.route('/removed')
def removed():
    dirs = os.listdir(path)
    # fileX = ['name', 'extension']
    file0 = [dirs[0].split('.', 1)[0], dirs[0].split('.', 1)[1]]
    file1 = [dirs[1].split('.', 1)[0], dirs[1].split('.', 1)[1]]
    if file0[1] == 'txt':
        txtfile = file0[0] + '.' + file0[1]
        xlsxfile = file1[0] + '.' + file1[1]
    else:
        txtfile = file1[0] + '.' + file1[1]
        xlsxfile = file0[0] + '.' + file0[1]
    filename = remove_columns(txtfile, xlsxfile)
    return render_template("col_rm/removed.html", filse=filename)


@bp.route("/download/<string:file_name>")
def download_file(file_name):
    try:
        return send_from_directory("col_rm_final", filename=file_name, as_attachment=True)
    except FileNotFoundError:
        abort(404)


def get_all_good_headers(txtfile):
    file = open(txtfile, 'r')
    dobre = file.readlines()
    good = []
    for i in dobre:
        good.append(i.strip())
        return good


def remove_columns(txtfile, xlsxfile):
    excel = path + '/' + xlsxfile
    txtfile = path + '/' + txtfile
    good = get_all_good_headers(txtfile)
    wbf = oxl.load_workbook(excel)
    sh0 = wbf[wbf.sheetnames[0]]

    number_of_columns = sh0.max_column

    i = 1
    while True:
        if i > number_of_columns:
            break
        cell = sh0.cell(row=1, column=i).value
        great = False
        for j in range(len(good)):
            header = good[j]
            great = True
            if header == cell:
                great = False
                i += 1
                break
        if great:
            sh0.delete_cols(i)
            number_of_columns -= 1
    finalname = xlsxfile.rsplit(".", 1)
    finale = finalname[0] + "_final.xlsx"
    wbf.save(path1 + '/' + finale)
    os.remove(txtfile)
    os.remove(excel)
    return finale
