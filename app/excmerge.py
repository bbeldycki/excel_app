from flask import Blueprint, request, flash, redirect, render_template, send_from_directory, abort, send_file
from werkzeug.utils import secure_filename
from pathlib import Path
import openpyxl as oxl
from copy import copy
import os

bpp = Blueprint('excelmrg', __name__)

allowed_extensions = set(['xlsx'])

path = "app/filesuploaded"
path1 = "app/excelmerged"
# for pythonanywhere paths
# path = "/home/Flippy9004/excel_app/app/static/filesuploaded"
# path1 = "/home/Flippy9004/excel_app/app/static/excelmerged"


def allowed_files(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


@bpp.route('/mrgupload', methods=["GET", "POST"])
def mrgupload():
    # Check if there is any "final" file in directory for final files
    # If there is, remove it
    dirs = os.listdir(path1)
    for file in dirs:
        os.remove(path1 + '/' + file)

    if request.method == "POST":

        if 'files[]' not in request.files:
            flash('No file part')
            return redirect(request.url)
        files = request.files.getlist('files[]')
        for file in files:
            if file.filename == '':
                flash('No file selected')
                return redirect(request.url)

        for file in files:
            if file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
                files.remove(file)

        if len(files) < 2:
            return render_template("exc_mrg/exc_upload.html", numfiles=len(files), wrong=True)

        filenames = []
        for file in files:
            if file and allowed_files(file.filename):
                filename = secure_filename(file.filename)
                filenames.append(filename)
                file.save(os.path.join(Path(path), filename))
        return render_template("exc_mrg/exc_upload.html", filse=filenames, lfse=len(filenames))
    return render_template("exc_mrg/exc_upload.html")


@bpp.route('/mrgremoved')
def mrgremoved():
    filename = merge_files()
    return render_template("exc_mrg/exc_removed.html", filse=filename)


@bpp.route("/mrgdownload/<string:file_name>")
def mrg_download_file(file_name):
    try:
        return send_from_directory("excelmerged", filename=file_name, as_attachment=True)
    except FileNotFoundError:
        abort(404)


def merge_files():
    xls_files_list = get_the_list_of_files_in_directory()
    wbf = oxl.load_workbook(path + '/' + xls_files_list[0])
    names = wbf.sheetnames
    # print('File ' + str(xls_files_list[0]) + ' copied. Done!')
    for i in range(1, len(xls_files_list)):
        wlx = oxl.load_workbook(path + '/' + xls_files_list[i])
        nameswlx = wlx.sheetnames
        nameswlx.remove('General Information')
        nameswlx.remove('LTBStatistics')
        for j in range(len(nameswlx)):
            sheet0 = wlx[nameswlx[j]]
            sheet1 = wbf[names[j]]
            copy_sheets(sheet0, sheet1)

    del wbf['General Information']
    fname = 'xlsxmerger.xlsx'
    wbf.save(path1 + '/' + fname)  # liczba wierszy w pierwszym pliku po dodawaniu
    for file in xls_files_list:
        os.remove(path + '/' + file)
    return fname


def get_the_list_of_files_in_directory():
    files_list = []
    dirs = os.listdir(path)
    for file in dirs:
        files_list.append(file)
    return files_list


def copy_sheets(sh0, sh1):
    rowfor0 = sh0.max_row
    rowfor1 = sh1.max_row
    if rowfor0 > 1:
        for i in range(2, sh0.max_row + 1):
            for j in range(1, sh0.max_column + 1):
                sh1.cell(row=rowfor1+i-1, column=j).value = sh0.cell(row=i, column=j).value
                if sh0.cell(row=i, column=j).has_style:
                    copy_prop(sh1.cell(row=rowfor1+i-1, column=j), sh0.cell(row=i, column=j))


def copy_prop(new_cell, old_cell):
    new_cell.data_type = copy(old_cell.data_type)
    new_cell.font = copy(old_cell.font)
    new_cell.border = copy(old_cell.border)
    new_cell.fill = copy(old_cell.fill)
    new_cell.number_format = copy(old_cell.number_format)
    new_cell.protection = copy(old_cell.protection)
    new_cell.alignment = copy(old_cell.alignment)
